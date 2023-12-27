#!/usr/bin/env python3

# Stanislas M. 2023-12-27

"""
usage: trend_api.py [-h] {computers}

List computers

positional arguments:
  {computers}  Action to perform

options:
  -h, --help   show this help message and exit
"""

import json
import requests
import argparse
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# disable ssl warning in case of proxy like Zscaler which breaks ssl...
requests.packages.urllib3.disable_warnings()

# Current date
now = datetime.now()
today = now.strftime("%Y-%m-%d-%H_%M_%S")

def export_to_csv(data):
    filename = today + "-trend-api-results.csv"
    with open(filename, "a") as f:
        f.write("computer,os,agent_status")
        for row in data:
            f.write(",".join(row) + "\n")

def export_to_excel(data):
    filename = today + "-trend-api-results.xlsx"
    wb = Workbook()
    ws = wb.active

    # Set file headers
    ws.append(['computer', 'os', 'agent_status'])

    # Define conditional formatting colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Apply conditional formatting
    for row in ws.iter_rows(min_row=2):
        # Agent column (column C)
        if row[2].value.startswith("warning"):
            row[2].fill = orange_fill
        elif row[2].value.startswith("error"):
            row[2].fill = red_fill
        elif row[2].value.startswith("active"):
            row[2].fill = green_fill

    # Add table with filters
    table = Table(displayName="ResultsTable", ref="A1:C{}".format(ws.max_row))
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Adjust columns width
    for column in ws.iter_cols():
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    print("Successfully created {}".format(filename))

# Read secrets from "secrets.json"
def read_secrets():
    with open('secrets.json') as f:
        secrets = json.load(f)
    if not (key in secrets for key in ['apiKey', 'proxyUrl']):
        raise ValueError('Error: Invalid secrets file')
    return secrets

# List Servers Trend Micro Agent status
def get_server_agent_status():

    secrets = read_secrets()

    headers = { 
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'api-version': 'v1',
        'Authorization': "ApiKey {}".format(secrets['apiKey'])
    }

    proxies = { 'http': secrets['proxyUrl'], 'https': secrets['proxyUrl'] }

    # Get enrolled computers
    url = "https://workload.us-1.cloudone.trendmicro.com/api/computers"
    response = requests.get(url, headers=headers, proxies=proxies, verify=False)
    json_response = json.loads(response.content)
    with open("computers.json", "w") as f:
        f.write(response.text)

    # f = open("computers.json", "r")
    # json_response = json.loads(f.read())
    # f.close()

    api_results = []
    cpt_api_results = 0
    cpt_active = 0
    cpt_warning = 0
    cpt_error = 0
    for computer in json_response["computers"]:
        # print("{},{},{}".format(computer['hostName'], computer['platform'], computer['computerStatus']['agentStatus']))

        if computer['computerStatus']['agentStatus'] == "active":
            cpt_active += 1
        
        elif computer['computerStatus']['agentStatus'] == "warning":
            cpt_warning += 1

        elif computer['computerStatus']['agentStatus'] == "error":
            cpt_error += 1

        cpt_api_results += 1

        api_results.append([computer['hostName'], computer['platform'], computer['computerStatus']['agentStatus']])

    print("{} agents are active, {} agents have warnings, {} agents have errors.".format(cpt_active, cpt_warning, cpt_error))
    print("Processed {} API results.".format(cpt_api_results))
    export_to_excel(api_results)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='List computers')
    parser.add_argument('action', choices=['computers'], help='Action to perform')
    args = parser.parse_args()
    try: 
        if args.action == 'computers':
            get_server_agent_status()

    except Exception as err:
        print("General error: " + str(err) + " check your secrets file.") 
        exit(1)
